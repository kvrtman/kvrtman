#!/usr/bin/env python3
"""Build Delivery_Scoreboard.xlsx, Kurt's private scoreboard.

Per the Jul call with Ivan: PMs are scored on delivery Speed (on time, or
faster) and Quality (expected level, minimal resyncs) against their seniority
bar. Initiative is logged as the growth dimension (not bonus). Calibration
lands on a single 1-5 score per person; first run is retroactive (past 3
months), not the December review.
"""
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SRC = 'Executive_Project_Tracker_v2.xlsx'
OUT = 'Delivery_Scoreboard.xlsx'

BLUE = '203AA9'
INK = '1A1A1A'
GRAY = '8A8A82'
FILLME = 'FFF3C9'

F_TITLE = Font(name='Arial', size=13, bold=True, color=BLUE)
F_HDR   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_BODY  = Font(name='Arial', size=10, color=INK)
F_BODYB = Font(name='Arial', size=10, bold=True, color=INK)
F_NOTE  = Font(name='Arial', size=9,  color=GRAY)
F_BLUEB = Font(name='Arial', size=10, bold=True, color=BLUE)
FILL_HDR = PatternFill('solid', fgColor=BLUE)
FILL_FILL = PatternFill('solid', fgColor=FILLME)
THIN_BOT = Border(bottom=Side(style='thin', color='D9D9D9'))
WRAP = Alignment(wrap_text=True, vertical='top')

OWNERS = ['JV Manlangit', 'Anton Betia', 'Manzil Balani', 'Elias Marcella', 'Jastin Lagumbay']
LOG = "'Delivery Log'"
# Delivery Log mirrors the tracker Master Rollup:
# A Owner, B Project, C Status, D Priority, E KPI, F Impact, G Added, H Due,
# I Delivered, J Speed, K Resyncs, L Summary, M Link

wb = openpyxl.Workbook()

# ------------------------------------------------------------- Scoreboard
ws = wb.active
ws.title = 'Scoreboard'
for col, w in {'A': 18, 'B': 15, 'C': 11, 'D': 9, 'E': 9, 'F': 8, 'G': 17, 'H': 13, 'I': 12}.items():
    ws.column_dimensions[col].width = w

ws['A1'] = 'Delivery Scoreboard, per owner. Kurt only.'
ws['A1'].font = F_TITLE
ws['A2'] = 'Private file: individual scores are not shared between team members. The shared tracker stays the team view of the work itself.'
ws['A2'].font = F_NOTE
ws['A3'] = 'Ivan, Jul call: was it delivered on time, yes, no, or faster; was it delivered at the level you expected. Resyncs are the quality signal. Computed from the Delivery Log tab.'
ws['A3'].font = F_NOTE

hdrs = ['Owner', 'Committed (Due set)', 'Delivered', 'Faster', 'On time', 'Late',
        'On time or faster', 'Overdue (open)', 'Avg resyncs']
for i, h in enumerate(hdrs, start=1):
    c = ws.cell(row=5, column=i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = Alignment(vertical='center', wrap_text=True)
ws.row_dimensions[5].height = 27

FIRST, LAST = 6, 6 + len(OWNERS) - 1
r = FIRST
for owner in OWNERS:
    m = f'({LOG}!$A$4:$A$400=$A{r})'
    ws.cell(row=r, column=1, value=owner).font = F_BODYB
    ws.cell(row=r, column=2, value=f'=SUMPRODUCT({m}*({LOG}!$H$4:$H$400<>""))')
    ws.cell(row=r, column=3, value=f'=SUMPRODUCT({m}*({LOG}!$I$4:$I$400<>""))')
    ws.cell(row=r, column=4, value=f'=SUMPRODUCT({m}*({LOG}!$I$4:$I$400<>"")*({LOG}!$H$4:$H$400<>"")*({LOG}!$I$4:$I$400<{LOG}!$H$4:$H$400))')
    ws.cell(row=r, column=5, value=f'=SUMPRODUCT({m}*({LOG}!$I$4:$I$400<>"")*({LOG}!$H$4:$H$400<>"")*({LOG}!$I$4:$I$400={LOG}!$H$4:$H$400))')
    ws.cell(row=r, column=6, value=f'=SUMPRODUCT({m}*({LOG}!$I$4:$I$400<>"")*({LOG}!$H$4:$H$400<>"")*({LOG}!$I$4:$I$400>{LOG}!$H$4:$H$400))')
    ws.cell(row=r, column=7, value=f'=IF(D{r}+E{r}+F{r}=0,"-",(D{r}+E{r})/(D{r}+E{r}+F{r}))')
    ws.cell(row=r, column=8, value=f'=SUMPRODUCT({m}*({LOG}!$H$4:$H$400<>"")*({LOG}!$I$4:$I$400="")*({LOG}!$C$4:$C$400<>"Completed")*({LOG}!$C$4:$C$400<>"Cancelled")*({LOG}!$H$4:$H$400<TODAY()))')
    ws.cell(row=r, column=9, value=f'=IFERROR(ROUND(AVERAGEIFS({LOG}!$K$4:$K$400,{LOG}!$A$4:$A$400,$A{r}),1),"-")')
    for col in range(1, 10):
        c = ws.cell(row=r, column=col)
        if col > 1:
            c.font = F_BODY
        c.border = THIN_BOT
    ws.cell(row=r, column=7).number_format = '0%'
    r += 1
ws.cell(row=r, column=1, value='Team').font = F_BLUEB
for col, letter in ((2, 'B'), (3, 'C'), (4, 'D'), (5, 'E'), (6, 'F'), (8, 'H')):
    c = ws.cell(row=r, column=col, value=f'=SUM({letter}{FIRST}:{letter}{LAST})'); c.font = F_BODYB
c = ws.cell(row=r, column=7, value=f'=IF(D{r}+E{r}+F{r}=0,"-",(D{r}+E{r})/(D{r}+E{r}+F{r}))')
c.font = F_BODYB; c.number_format = '0%'

r += 2
for line in [
    'How to read it: Committed = rows with an agreed Due date. Speed needs Due dates, so set them at standup; a delivery without a Due date cannot be scored.',
    'Refresh weekly: copy rows A4 down to M from the tracker\'s Master Rollup and paste into the Delivery Log tab (same columns). In Google Sheets, the Delivery Log tab can be replaced with IMPORTRANGE for auto-sync.',
    'Feeds the quarterly Revolut-style scorecard: Speed from on time or faster, Quality from resyncs and level delivered, Complexity judged per project. Deliverables = (Speed + Quality) x Complexity.',
]:
    ws.cell(row=r, column=1, value=line).font = F_NOTE
    r += 1

# ------------------------------------------------------------- Delivery Log
src = openpyxl.load_workbook(SRC)
sm = src['Master Rollup']
ws2 = wb.create_sheet('Delivery Log')
for col, w in {'A': 16, 'B': 30, 'C': 13, 'D': 11, 'E': 26, 'F': 34, 'G': 9, 'H': 9,
               'I': 9.5, 'J': 9.5, 'K': 9, 'L': 50, 'M': 46}.items():
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = 'A4'
ws2['A1'] = 'Mirror of the tracker Master Rollup (columns A to M). Refresh by paste, or IMPORTRANGE in Google Sheets. The Scoreboard reads Owner (A), Status (C), Due (H), Delivered (I), Resyncs (K).'
ws2['A1'].font = F_NOTE
for i, h in enumerate(['Owner', 'Project', 'Status', 'Priority', 'KPI Key Driver', 'Impact (bottom-up)',
                       'Added', 'Due', 'Delivered', 'Speed', 'Resyncs', 'Summary', 'ClickUp / BRD Link'], start=1):
    c = ws2.cell(row=3, column=i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = Alignment(vertical='center')
ws2.row_dimensions[3].height = 21.75
r = 4
for row in sm.iter_rows(min_row=4, max_col=13):
    if row[0].value is None:
        continue
    for i, c in enumerate(row, start=1):
        v = c.value
        if i == 10 and isinstance(v, str) and v.startswith('='):
            v = f'=IF(I{r}="","",IF(H{r}="","-",IF(I{r}<H{r},"Faster",IF(I{r}=H{r},"On time","Late"))))'
        nc = ws2.cell(row=r, column=i, value=v)
        nc.font = F_BODY
        nc.border = THIN_BOT
        nc.alignment = Alignment(vertical='top')
        if i in (7, 8, 9):
            nc.number_format = 'dd mmm'
    ws2.cell(row=r, column=1).font = F_BODYB
    r += 1

# ------------------------------------------------------------- Initiative Log
ws3 = wb.create_sheet('Initiative Log')
for col, w in {'A': 9, 'B': 16, 'C': 58, 'D': 10, 'E': 10, 'F': 14, 'G': 12, 'H': 40}.items():
    ws3.column_dimensions[col].width = w
ws3['A1'] = 'Initiative Log: the growth dimension (not bonus)'
ws3['A1'].font = F_TITLE
ws3['A2'] = 'Ivan, Jul call: you cannot grow if you do not do initiatives. Track who connects the numbers to actions. Assess each idea easy vs difficult and high vs low impact; it gets its own priority; then run it.'
ws3['A2'].font = F_NOTE
hdr = ['Date', 'Person', 'Initiative (problem, proposed action)', 'Effort', 'Impact', 'Priority (auto)', 'Status', 'Notes']
for i, h in enumerate(hdr, start=1):
    c = ws3.cell(row=4, column=i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = Alignment(vertical='center')
ws3.row_dimensions[4].height = 21.75
# seeded with the real item from the call
ws3['A5'] = dt.date(2026, 7, 20); ws3['A5'].number_format = 'dd mmm'
ws3['B5'] = 'Anton Betia'
ws3['C5'] = 'Deals listing sorts P9,999 item first while the popular P1,000 item is last; add cheapest-first sorting'
ws3['D5'] = 'Easy'
ws3['E5'] = 'High'
ws3['G5'] = 'In tracker'
ws3['H5'] = 'Raised by Ivan on the Jul call; Kurt committed the fix'
for rr in range(5, 40):
    ws3.cell(row=rr, column=6, value=f'=IF(OR(D{rr}="",E{rr}=""),"",IF(AND(D{rr}="Easy",E{rr}="High"),"P1 do now",IF(E{rr}="High","P2 plan",IF(D{rr}="Easy","P3 quick win","P4 later"))))')
    for col in range(1, 9):
        c = ws3.cell(row=rr, column=col)
        c.font = F_BODY; c.border = THIN_BOT; c.alignment = WRAP
    ws3.cell(row=rr, column=1).number_format = 'dd mmm'
ws3['B5'].font = F_BODYB
dv_e = DataValidation(type='list', formula1='"Easy,Medium,Hard"', allow_blank=True)
dv_i = DataValidation(type='list', formula1='"High,Medium,Low"', allow_blank=True)
dv_s = DataValidation(type='list', formula1='"Proposed,Approved,In tracker,Dropped"', allow_blank=True)
ws3.add_data_validation(dv_e); ws3.add_data_validation(dv_i); ws3.add_data_validation(dv_s)
dv_e.add('D5:D39'); dv_i.add('E5:E39'); dv_s.add('G5:G39')

# ------------------------------------------------------------- Calibration
ws4 = wb.create_sheet('Calibration')
for col, w in {'A': 18, 'B': 20, 'C': 15, 'D': 11, 'E': 11, 'F': 20, 'G': 13, 'H': 46}.items():
    ws4.column_dimensions[col].width = w
ws4['A1'] = 'Calibration: one score per person (5 to 1)'
ws4['A1'].font = F_TITLE
ws4['A2'] = 'Run NOW for the past 3 months; do not wait for the December review (Ivan, Jul call). Compare the result to your gut feel; when it corresponds, adopt it. How you arrive at the score, no one cares; the score is what calibrates.'
ws4['A2'].font = F_NOTE
ws4['A3'] = 'The bar is the person, not the peer: "I am not comparing you with him. I am comparing your bar with your performance." Juniors get a mistake discount; non-juniors must arrive fully thought through.'
ws4['A3'].font = F_NOTE
hdr = ['Person', 'Seniority (the bar)', 'On time or faster', 'Avg resyncs', 'Initiatives', 'Scorecard overall grade', 'Score (5-1)', 'Notes / evidence']
for i, h in enumerate(hdr, start=1):
    c = ws4.cell(row=5, column=i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = Alignment(vertical='center', wrap_text=True)
ws4.row_dimensions[5].height = 27
CAL = [
    ('JV Manlangit', ''),
    ('Anton Betia', ''),
    ('Manzil Balani', ''),
    ('Elias Marcella', 'Senior'),
    ('Jastin Lagumbay', ''),
]
r = 6
for person, senior in CAL:
    ws4.cell(row=r, column=1, value=person).font = F_BODYB
    sc = ws4.cell(row=r, column=2, value=senior if senior else None)
    sc.font = F_BODY
    if not senior:
        sc.fill = FILL_FILL
    ws4.cell(row=r, column=3, value=f'=IFERROR(INDEX(Scoreboard!$G${FIRST}:$G${LAST},MATCH($A{r},Scoreboard!$A${FIRST}:$A${LAST},0)),"-")').number_format = '0%'
    ws4.cell(row=r, column=4, value=f'=IFERROR(INDEX(Scoreboard!$I${FIRST}:$I${LAST},MATCH($A{r},Scoreboard!$A${FIRST}:$A${LAST},0)),"-")')
    ws4.cell(row=r, column=5, value=f'=COUNTIFS(\'Initiative Log\'!$B$5:$B$39,$A{r})')
    g = ws4.cell(row=r, column=6); g.fill = FILL_FILL
    s = ws4.cell(row=r, column=7); s.fill = FILL_FILL
    ws4.cell(row=r, column=8).fill = FILL_FILL
    for col in range(1, 9):
        c = ws4.cell(row=r, column=col)
        if col > 2:
            c.font = F_BODY
        c.border = THIN_BOT
    r += 1
dv_sen = DataValidation(type='list', formula1='"Junior,Mid,Senior,Lead"', allow_blank=True)
dv_grade = DataValidation(type='list', formula1='"A-player,Above the Bar,Underperformer"', allow_blank=True)
dv_score = DataValidation(type='list', formula1='"5,4,3,2,1"', allow_blank=True)
ws4.add_data_validation(dv_sen); ws4.add_data_validation(dv_grade); ws4.add_data_validation(dv_score)
dv_sen.add(f'B6:B{r-1}'); dv_grade.add(f'F6:F{r-1}'); dv_score.add(f'G6:G{r-1}')
r += 1
ws4.cell(row=r, column=1, value='Seniority for Elias per Ivan: superstar bar, well paid, the only discount is domain knowledge. Set the rest of the bars yourself and confirm with Ivan.').font = F_NOTE
r += 1
ws4.cell(row=r, column=1, value='Scorecard overall grade comes from each person\'s QuantumLight Product Owner scorecard (one copy per person per quarter). Yellow = fill by hand.').font = F_NOTE

wb.save(OUT)
print('saved', OUT)
