#!/usr/bin/env python3
"""Card bureau decision one-pager (IDEMIA vs MM), requested by Ivan on the Jul call.

Rule agreed: compare apples to apples on the full scope, go with the cheaper.
Yellow cells = prices to fill from the two quotes. Totals, difference and the
recommendation compute themselves.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = 'Card_Bureau_Decision_OnePager.xlsx'

BLUE = '203AA9'
INK = '1A1A1A'
GRAY = '8A8A82'
FILLME = 'FFF3C9'

F_TITLE = Font(name='Arial', size=13, bold=True, color=BLUE)
F_HDR   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_BODY  = Font(name='Arial', size=10, color=INK)
F_BODYB = Font(name='Arial', size=10, bold=True, color=INK)
F_NOTE  = Font(name='Arial', size=9,  color=GRAY)
F_BLUEB = Font(name='Arial', size=10, bold=True, color=BLUE)
FILL_HDR = PatternFill('solid', fgColor=BLUE)
FILL_FILL = PatternFill('solid', fgColor=FILLME)
THIN_BOT = Border(bottom=Side(style='thin', color='D9D9D9'))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Card Bureau Decision'
for col, w in {'A': 34, 'B': 14, 'C': 14, 'D': 50}.items():
    ws.column_dimensions[col].width = w

ws['A1'] = 'First card order: bureau selection (IDEMIA vs MM)'
ws['A1'].font = F_TITLE
ws['A2'] = 'Rule agreed with Ivan: like-for-like scope, all lines in (printing, personalization, envelope, collaterals, QR personalization), go with the cheaper. Artwork is paid either way at market price, no hard feelings.'
ws['A2'].font = F_NOTE
ws['A3'] = 'Context: MM appears to add margin on top of bureau production (no volume discount passed through), so a direct like-for-like total decides. Yellow = fill from the two final quotes.'
ws['A3'].font = F_NOTE

ws['A5'] = 'Volume assumption (cards in first order)'
ws['A5'].font = F_BODYB
ws['B5'].fill = FILL_FILL
ws['B5'].number_format = '#,##0'
ws['D5'] = 'Set the same volume for both quotes; bureau pricing is volume-driven.'
ws['D5'].font = F_NOTE

hdr_row = 7
for i, h in enumerate(['Cost line (per card, PHP)', 'IDEMIA', 'MM', 'Notes'], start=1):
    c = ws.cell(row=hdr_row, column=i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = Alignment(vertical='center')
ws.row_dimensions[hdr_row].height = 21.75

per_card = [
    'Card printing (numberless body)',
    'Personalization',
    'QR code personalization',
    'Envelope',
    'Collaterals / inserts',
    'Delivery and logistics',
    'Other per-card fees',
]
r = hdr_row + 1
first_pc = r
for line in per_card:
    ws.cell(row=r, column=1, value=line).font = F_BODY
    for col in (2, 3):
        c = ws.cell(row=r, column=col); c.fill = FILL_FILL; c.number_format = '#,##0.00'
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = THIN_BOT
    r += 1
last_pc = r - 1

ws.cell(row=r, column=1, value='Per-card total').font = F_BODYB
ws.cell(row=r, column=2, value=f'=SUM(B{first_pc}:B{last_pc})').font = F_BODYB
ws.cell(row=r, column=3, value=f'=SUM(C{first_pc}:C{last_pc})').font = F_BODYB
for col in (2, 3):
    ws.cell(row=r, column=col).number_format = '#,##0.00'
for col in range(1, 5):
    ws.cell(row=r, column=col).border = THIN_BOT
pc_total = r
r += 2

for i, h in enumerate(['One-time items (PHP)', 'IDEMIA', 'MM', 'Notes'], start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = F_HDR; c.fill = FILL_HDR; c.alignment = Alignment(vertical='center')
ws.row_dimensions[r].height = 21.75
r += 1
first_ot = r
for line, note_txt in (('Artwork', 'Market price either way; already scoped by Nelly incl. artwork.'),
                       ('Setup / tooling', ''),
                       ('Other one-time fees', '')):
    ws.cell(row=r, column=1, value=line).font = F_BODY
    for col in (2, 3):
        c = ws.cell(row=r, column=col); c.fill = FILL_FILL; c.number_format = '#,##0'
    ws.cell(row=r, column=4, value=note_txt if note_txt else None).font = F_NOTE
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = THIN_BOT
    r += 1
last_ot = r - 1
ws.cell(row=r, column=1, value='One-time total').font = F_BODYB
ws.cell(row=r, column=2, value=f'=SUM(B{first_ot}:B{last_ot})').font = F_BODYB
ws.cell(row=r, column=3, value=f'=SUM(C{first_ot}:C{last_ot})').font = F_BODYB
for col in (2, 3):
    ws.cell(row=r, column=col).number_format = '#,##0'
for col in range(1, 5):
    ws.cell(row=r, column=col).border = THIN_BOT
ot_total = r
r += 2

ws.cell(row=r, column=1, value='Total for the first order').font = F_BLUEB
ws.cell(row=r, column=2, value=f'=IF($B$5="","fill volume",B{pc_total}*$B$5+B{ot_total})').font = F_BODYB
ws.cell(row=r, column=3, value=f'=IF($B$5="","fill volume",C{pc_total}*$B$5+C{ot_total})').font = F_BODYB
for col in (2, 3):
    ws.cell(row=r, column=col).number_format = '#,##0'
grand = r
r += 1
ws.cell(row=r, column=1, value='Difference (savings by going cheaper)').font = F_BODYB
ws.cell(row=r, column=2, value=f'=IF($B$5="","-",ABS(B{grand}-C{grand}))').font = F_BODYB
ws.cell(row=r, column=2).number_format = '#,##0'
r += 1
ws.cell(row=r, column=1, value='Recommendation').font = F_BLUEB
rec = ws.cell(row=r, column=2, value=f'=IF($B$5="","-",IF(B{grand}<C{grand},"IDEMIA",IF(C{grand}<B{grand},"MM","Tie")))')
rec.font = Font(name='Arial', size=11, bold=True, color=BLUE)
r += 2
for line in [
    'Decision logic: same scope on every line, same volume, cheaper total wins. If MM cannot beat direct bureau pricing, there is no value added in the middle.',
    'Outcome statement for the share-out: we compared like for like, we go with [winner], and we save PHP [difference] on the first order.',
]:
    ws.cell(row=r, column=1, value=line).font = F_NOTE
    r += 1

ws.print_area = f'A1:D{r}'
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

wb.save(OUT)
print('saved', OUT)
