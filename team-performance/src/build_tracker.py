#!/usr/bin/env python3
"""Build Executive_Project_Tracker_v2.xlsx from the original tracker.

Shared with the whole team. Reflects the Jul call with Ivan: PMs are scored on
delivery (Speed x Quality vs the seniority bar), not on hitting line volumes.
Volumes are management-owned context. Initiative is the growth dimension.
Adds Added / Due / Delivered / Speed (auto) / Resyncs columns.
"""
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SRC = 'tracker.xlsx'
OUT = 'Executive_Project_Tracker_v2.xlsx'

BLUE = '203AA9'
INK = '1A1A1A'
GRAY = '8A8A82'
BAND = 'F6F4EF'
FILLME = 'FFF3C9'

STATUS_STYLE = {
    'Blocked':     ('C0392B', True, 'FBEAE7'),
    'Cancelled':   ('8A8A82', True, 'ECE8E1'),
    'Completed':   ('1E8E5A', True, 'E7F4EE'),
    'In Progress': ('203AA9', True, 'ECEEF8'),
    'Not Started': ('8A8A82', True, 'F2EEE7'),
}

F_TITLE = Font(name='Arial', size=13, bold=True, color=BLUE)
F_HDR   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_BODY  = Font(name='Arial', size=10, color=INK)
F_BODYB = Font(name='Arial', size=10, bold=True, color=INK)
F_NOTE  = Font(name='Arial', size=9,  color=GRAY)
F_BLUEB = Font(name='Arial', size=10, bold=True, color=BLUE)
FILL_HDR  = PatternFill('solid', fgColor=BLUE)
FILL_BAND = PatternFill('solid', fgColor=BAND)
FILL_FILL = PatternFill('solid', fgColor=FILLME)
THIN_BOT = Border(bottom=Side(style='thin', color='D9D9D9'))
WRAP = Alignment(wrap_text=True, vertical='top')
TOP = Alignment(vertical='top')

OWNER_TABS = {'JV': 'JV Manlangit', 'Anton': 'Anton Betia', 'Manzil': 'Manzil Balani',
              'Elias': 'Elias Marcella', 'Jastin': 'Jastin Lagumbay'}

IMPACTS = {
    ('JV', 'Motorcycle Loans (KServico)'): 'Moto monthly run rate, gated by Nhat',
    ('Anton', 'Special Deals Offer to Low-Risk Customers'): 'Conversion lift on low-risk base, size with Nhat before comms',
    ('Manzil', 'Deals Merchant BD Pipeline'): 'New merchants live and disbursing, ticket uplift',
    ('Elias', 'Credit Line (Launch)'): 'CL utilization and drawdown, target set ~30 days post-launch',
    ('Elias', 'Access Card'): 'Per-card break-even on interchange at target CAC',
    ('Jastin', 'QRPh Chargeback Escalation (Netbank/AUB)'): 'Protects the QRPh run rate',
    ('Jastin', 'Integrate Merchant Guide to Billease Docs'): 'Fewer onboarding tickets via self-serve guide',
}
DUES = {('Anton', 'Special Deals Offer to Low-Risk Customers'): dt.date(2026, 7, 20)}
DELIVERED = {
    ('JV', 'SAP Announcement & Quiz Module v1'): dt.date(2026, 7, 3),
    ('JV', 'Credit Line - Agent Announcement (SAP)'): dt.date(2026, 7, 16),
    ('Manzil', 'Maxicare PRIMA Launch on Deals'): dt.date(2026, 7, 10),
}

src = openpyxl.load_workbook(SRC)
wb = openpyxl.Workbook()
wb.remove(wb.active)

def title(ws, cell, text):
    ws[cell] = text; ws[cell].font = F_TITLE

def note(ws, cell, text):
    ws[cell] = text; ws[cell].font = F_NOTE

def hdr_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = F_HDR; c.fill = FILL_HDR; c.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 21.75

def style_status(cell):
    s = STATUS_STYLE.get(cell.value)
    if s:
        cell.font = Font(name='Arial', size=10, bold=s[1], color=s[0])
        cell.fill = PatternFill('solid', fgColor=s[2])

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ---------------------------------------------------------------- README
ws = wb.create_sheet('README')
set_widths(ws, {'A': 2.5, 'B': 112})
L = [
    ('T', 'Executive Project Tracker v2, POS Product Team (H2 2026)'),
    ('B', 'One file, shared with the whole team. One tab per owner plus a consolidated Master. Linked to the PBT Metric Tree.'),
    ('S', ''),
    ('H', 'How the team is measured (agreed with Ivan, Jul call)'),
    ('B', 'PMs are not scored on hitting line volumes. Volume and PBT numbers are management-owned; the team is not punished for them ("they are not magicians"). Sales-style volume KPIs belong to the network team; our target is profit, and sometimes less sales means more profit.'),
    ('B', 'What PMs own and are scored on is DELIVERY of the agreed projects: Speed (on time, or faster) and Quality (delivered at the expected level, thought through, minimal resyncs). The bar is the person\'s seniority: juniors get a mistake discount, non-juniors must deliver magic. "I am not comparing you with him. I am comparing your bar with your performance."'),
    ('B', 'Initiative is the third dimension, assessed for growth, not for bonus: does the person know their numbers cold and connect problems to actions, bringing prioritized initiatives (easy vs difficult, high vs low impact).'),
    ('S', ''),
    ('H', 'Purpose'),
    ('B', 'One row per project, straight to the point. ClickUp stays the main working document and source of truth.'),
    ('B', 'This sheet is the executive-level view: what the project is, whether it is moving, what it is expected to move (Impact), and whether we delivered when we said we would.'),
    ('B', 'Everyone sees everything here. The point is a single, centralized view of what the team is doing.'),
    ('S', ''),
    ('H', 'The columns'),
    ('B', 'Project, Status, Priority, KPI Key Driver, Impact (bottom-up), Added, Due, Delivered, Speed, Resyncs, Summary, ClickUp or BRD Link. The Master also has an Owner column.'),
    ('B', 'Added = the date the project entered the tracker. Due = the date agreed with the owner. Delivered = the date it actually shipped.'),
    ('B', 'Speed is automatic from Due vs Delivered: Faster, On time, or Late. A dash means it was delivered without an agreed Due date; set Due dates so Speed can be scored.'),
    ('B', 'Resyncs counts how many times the work bounced back for rework or another sync. It is the quality proxy: a well-thought-through delivery needs no repeated syncs (Ivan: "if I have to resync and come back again and again, the quality goes down automatically").'),
    ('B', 'Status and Priority are dropdowns. Yellow cells are blanks the owner must fill: Impact, Added and Due before a project enters In Progress; Delivered when it ships.'),
    ('B', 'In Google Sheets, Added and Delivered can be stamped automatically with a small onEdit script (see team-performance/src in the repo), or with iterative-calculation formulas.'),
    ('S', ''),
    ('H', 'Status and Priority values'),
    ('B', 'Status: Not Started, In Progress, Blocked, Completed, Cancelled.'),
    ('B', 'Priority: Urgent, High, Normal, Low.'),
    ('S', ''),
    ('H', 'Ranking rule'),
    ('B', 'Rows are ordered Urgent, High, Normal, Low, with Completed and Cancelled sent to the bottom.'),
    ('S', ''),
    ('H', 'Operating cadence'),
    ('B', 'Weekly standup: walk the Master top-down, blocked and overdue first, update Due, Delivered and Resyncs. Details live in ClickUp via the link column.'),
    ('B', 'Monthly: KPI readout on the KPI Targets H2 tab against Nhat\'s PBT-per-line model and Jenny\'s metric-tree dashboard. Context for everyone; scoring for no one.'),
    ('B', 'Quarterly: individual reviews on the QuantumLight (Revolut) Product Owner scorecard, graded against the seniority talent bar. Speed and Resyncs here are the evidence. Kurt runs a retro review for the past 3 months now rather than waiting for December (Ivan, Jul call).'),
    ('S', ''),
    ('H', 'Lock each tab to one person (done inside Google Sheets)'),
    ('B', '1. Share the whole file with the 5 owners as Editor, and the wider team as Viewer.'),
    ('B', '2. On each owner tab, use Data, then Protect sheets and ranges, then Set permissions, then Custom.'),
    ('B', '3. Add only that owner\'s email plus your own, then Done.'),
    ('S', ''),
    ('H', 'Design'),
    ('B', 'Minimalist, per the Billease brand: white-dominant, structural blue 203AA9 headers, ink text, rationed color.'),
]
r = 2
for kind, text in L:
    if kind == 'S':
        r += 1; continue
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = F_TITLE if kind == 'T' else (F_BLUEB if kind == 'H' else F_BODY)
    r += 1

# ------------------------------------------------- owner tabs + master data
def enrich(tab, rows):
    for d in rows:
        key = (tab, d['project'])
        d['impact'] = IMPACTS.get(key, '')
        d['due'] = DUES.get(key, '')
        d['delivered'] = DELIVERED.get(key, '')
    return rows

owner_rows = {}
for tab in OWNER_TABS:
    s = src[tab]
    rows = []
    for row in s.iter_rows(min_row=4):
        if row[0].value is None:
            continue
        rows.append(dict(project=row[0].value, status=row[1].value, prio=row[2].value,
                         kpi=row[3].value, summary=row[4].value,
                         link=row[5].value if len(row) > 5 else None))
    owner_rows[tab] = enrich(tab, rows)

HDRS = ['Project', 'Status', 'Priority', 'KPI Key Driver', 'Impact (bottom-up)',
        'Added', 'Due', 'Delivered', 'Speed', 'Resyncs', 'Summary', 'ClickUp / BRD Link']
OWNER_WIDTHS = {'A': 32, 'B': 13, 'C': 11, 'D': 26, 'E': 34, 'F': 9, 'G': 9, 'H': 9.5,
                'I': 9.5, 'J': 9, 'K': 50, 'L': 46}

def write_project_row(ws, r, vals, status_col, banded):
    status = vals[status_col - 1]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v if v != '' else None)
        c.font = F_BODY; c.border = THIN_BOT
        c.alignment = TOP
        if banded:
            c.fill = FILL_BAND
    ws.cell(row=r, column=1).font = F_BODYB
    style_status(ws.cell(row=r, column=status_col))
    return status

def yellow_blanks(ws, r, status, cols):
    for col, kind in cols.items():
        c = ws[f'{col}{r}']
        if c.value is None:
            if (kind == 'open' and status not in ('Completed', 'Cancelled')) or \
               (kind == 'all' and status != 'Cancelled') or \
               (kind == 'done' and status == 'Completed'):
                c.fill = FILL_FILL

for tab, full in OWNER_TABS.items():
    ws = wb.create_sheet(tab)
    set_widths(ws, OWNER_WIDTHS)
    ws.freeze_panes = 'A4'
    ws['A1'] = f'{tab}: edit only this tab. Yellow = fill me (Impact, Added and Due before In Progress; Delivered when shipped; update Resyncs when work bounces back).'
    ws['A1'].font = F_NOTE
    hdr_row(ws, 3, HDRS)
    r = 4
    for i, d in enumerate(owner_rows[tab]):
        speed = f'=IF(H{r}="","",IF(G{r}="","-",IF(H{r}<G{r},"Faster",IF(H{r}=G{r},"On time","Late"))))'
        vals = [d['project'], d['status'], d['prio'], d['kpi'], d['impact'],
                '', d['due'], d['delivered'], speed, '', d['summary'], d['link']]
        write_project_row(ws, r, vals, status_col=2, banded=(i % 2 == 1))
        for col in ('F', 'G', 'H'):
            ws[f'{col}{r}'].number_format = 'dd mmm'
        yellow_blanks(ws, r, d['status'], {'E': 'open', 'F': 'all', 'G': 'open', 'H': 'done'})
        r += 1
    dv1 = DataValidation(type='list', formula1='"Not Started,In Progress,Blocked,Completed,Cancelled"', allow_blank=True)
    dv2 = DataValidation(type='list', formula1='"Urgent,High,Normal,Low"', allow_blank=True)
    ws.add_data_validation(dv1); ws.add_data_validation(dv2)
    dv1.add('B4:B400'); dv2.add('C4:C400')

# Master Rollup
ws = wb.create_sheet('Master Rollup', 1)
set_widths(ws, {'A': 16, 'B': 30, 'C': 13, 'D': 11, 'E': 26, 'F': 34, 'G': 9, 'H': 9,
                'I': 9.5, 'J': 9.5, 'K': 9, 'L': 50, 'M': 46})
ws.freeze_panes = 'A4'
ws['A1'] = 'Master Rollup: all projects, ranked (Urgent, High, Normal, Low, done at bottom). Team = Viewer. Yellow = owner to fill. Speed is automatic.'
ws['A1'].font = F_NOTE
hdr_row(ws, 3, ['Owner'] + HDRS)
sm = src['Master Rollup']
full_to_tab = {v: k for k, v in OWNER_TABS.items()}
r = 4
i = 0
for row in sm.iter_rows(min_row=4):
    if row[0].value is None:
        continue
    owner, project, status, prio, kpi, summary = (row[0].value, row[1].value, row[2].value,
                                                  row[3].value, row[4].value, row[5].value)
    link = row[6].value if len(row) > 6 else None
    tab = full_to_tab[owner]
    key = (tab, project)
    speed = f'=IF(I{r}="","",IF(H{r}="","-",IF(I{r}<H{r},"Faster",IF(I{r}=H{r},"On time","Late"))))'
    vals = [owner, project, status, prio, kpi, IMPACTS.get(key, ''), '', DUES.get(key, ''),
            DELIVERED.get(key, ''), speed, '', summary, link]
    write_project_row(ws, r, vals, status_col=3, banded=(i % 2 == 1))
    for col in ('G', 'H', 'I'):
        ws[f'{col}{r}'].number_format = 'dd mmm'
    yellow_blanks(ws, r, status, {'F': 'open', 'G': 'all', 'H': 'open', 'I': 'done'})
    r += 1; i += 1
dv1 = DataValidation(type='list', formula1='"Not Started,In Progress,Blocked,Completed,Cancelled"', allow_blank=True)
dv2 = DataValidation(type='list', formula1='"Urgent,High,Normal,Low"', allow_blank=True)
ws.add_data_validation(dv1); ws.add_data_validation(dv2)
dv1.add('C4:C400'); dv2.add('D4:D400')

# ------------------------------------------------------- KPI Targets H2
ws = wb.create_sheet('KPI Targets H2', 2)
set_widths(ws, {'A': 26, 'B': 12, 'C': 16, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 'H': 8, 'I': 8, 'J': 10, 'K': 46})
title(ws, 'A1', 'KPI Targets H2 2026, per line (Tier 0 / Tier 1). Management-owned context.')
note(ws, 'A2', 'These numbers are owned by management (Ivan + Kurt, fed by Nhat\'s model and Jenny\'s dashboard). The team is NOT scored or punished on these outcomes; every PM must know their line cold and connect problems to actions. PM scoring lives in delivery (Speed x Quality on the tracker). Yellow = set with Ivan or fill monthly.')
MONTHS = ['Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

def targets_block(ws, r0, section, rows, fmt='#,##0'):
    ws.cell(row=r0, column=1, value=section).font = F_BLUEB
    hdr_row(ws, r0 + 1, ['Line', 'Phase', 'PM (context)'] + MONTHS + ['YTD att.', 'Basis / notes'])
    r = r0 + 2
    for line, phase, owner, notes in rows:
        t, a = r, r + 1
        ws.cell(row=t, column=1, value=line).font = F_BODYB
        ws.cell(row=t, column=2, value=phase).font = F_BODY
        ws.cell(row=t, column=3, value=owner).font = F_BODY
        ws.cell(row=a, column=1, value='   actual').font = F_NOTE
        for j in range(6):
            ct = ws.cell(row=t, column=4 + j)
            ca = ws.cell(row=a, column=4 + j)
            ct.number_format = fmt; ca.number_format = fmt
            ct.fill = FILL_FILL; ca.fill = FILL_FILL
        att = ws.cell(row=a, column=10,
                      value=f'=IFERROR(IF(COUNT(D{a}:I{a})=0,"-",SUM(D{a}:I{a})/SUMPRODUCT((D{a}:I{a}<>"")*D{t}:I{t})),"-")')
        att.number_format = '0%'; att.font = F_BODYB
        ws.cell(row=t, column=11, value=notes).font = F_NOTE
        ws.cell(row=t, column=11).alignment = WRAP
        for row_i in (t, a):
            for col in range(1, 12):
                ws.cell(row=row_i, column=col).border = THIN_BOT
        r += 2
    return r + 1

r = targets_block(ws, 4, 'Line outcomes, monthly volume (PHP M disbursement)', [
    ('QRPh disbursement', 'Star', 'Jastin', 'Run-rate to confirm with Ivan. Hold and defend the star line (chargebacks, onboarding cycle time).'),
    ('Deals disbursement', 'Question', 'Anton + Manzil', 'Staged monthly ramp at positive NPM. Stage sizes and timing set by management; the reject-salvage plan sizes the ramp.'),
    ('Moto (KServico)', 'Question', 'JV', 'Run-rate gated by Nhat\'s unit-economics check.'),
    ('Deals, managed accounts', 'Question', 'Manzil', 'ProTech, Maxicare, 3Cat. Maxicare live: soft launch first week, marketing on this week.'),
])
r = targets_block(ws, r, 'Credit Line (utilization %, target set ~30 days post-launch)', [
    ('CL utilization / drawdown', 'New launch', 'Elias', 'Launched 15 Jul (F&F/ETB). Baseline first, then set the target with Ivan.'),
], fmt='0%')
r = targets_block(ws, r, 'Tier 0, PBT per line (PHP M, from Nhat\'s model)', [
    ('QRPh', 'Star', 'w/ Nhat', 'From Nhat\'s model once live.'),
    ('Online', 'Maintain', 'w/ Nhat', 'From Nhat\'s model once live.'),
    ('Moto / Deals', 'Question', 'w/ Nhat', 'Phase says push volume, watch NPM.'),
    ('Credit Line', 'New launch', 'w/ Nhat', 'Model rows live post-launch.'),
    ('Solar', 'Question', 'w/ Nhat', 'Kill or keep decision pending.'),
    ('Offline', 'Cash cow', 'w/ Nhat', 'Maintain, no PM by design.'),
])
ws.cell(row=r, column=1, value='Milestones (dates, not monthly numbers)').font = F_BLUEB
hdr_row(ws, r + 1, ['Milestone', 'Owner', 'Due', 'Done', 'Notes', '', '', '', '', '', ''])
mile = [
    ('CL: Gorg comment list turned into a checklist, validated 1:1', 'Elias + Kurt', dt.date(2026, 7, 22), 'Every item closed before scale-up. Do not abandon the child (Ivan).'),
    ('CL: iOS bugs fixed and apps submitted (Renata)', 'Elias', dt.date(2026, 7, 21), '6 iOS bugs found; fix committed same day.'),
    ('Card: bureau decision one-pager (IDEMIA vs MM, like for like)', 'Kurt + Elias', dt.date(2026, 7, 24), 'Rule: apples to apples, go cheaper. Share transparently.'),
    ('Card: Amar meeting, raise 60-day clause carefully', 'Ivan + Kurt', dt.date(2026, 7, 24), 'Signed 15 Jun. Some of their integrations were late too; do not be harsh.'),
    ('Deals: cheapest-first sorting on listing', 'Anton', dt.date(2026, 7, 24), 'From Ivan: P9,999 shows first, popular P1,000 last today.'),
    ('Deals: Maxicare marketing plans switched on', 'Anton + Manzil', dt.date(2026, 7, 24), 'Internal comms + PR aimed at investors and future Deals partners.'),
    ('Retro scorecard run, past 3 months, all 5 PMs', 'Kurt', dt.date(2026, 7, 24), 'Do not wait for December (Ivan). Calibrate to a 1-5 score.'),
    ('Solar: kill-or-keep decision', 'Anton', None, 'Per KPI tab D3.'),
    ('CL utilization target set with Ivan', 'Elias + Kurt', dt.date(2026, 8, 15), '30 days post-launch.'),
    ('PBT-per-line model live, monthly', 'Nhat', None, 'Reconciling to actuals.'),
    ('Metric-tree dashboard (Tier 1-3)', 'Jenny', None, 'One source for this tab.'),
]
r2 = r + 2
for m, o, due, n in mile:
    ws.cell(row=r2, column=1, value=m).font = F_BODYB
    ws.cell(row=r2, column=1).alignment = WRAP
    ws.cell(row=r2, column=2, value=o).font = F_BODY
    cd = ws.cell(row=r2, column=3, value=due); cd.number_format = 'dd mmm'; cd.font = F_BODY
    if due is None:
        cd.fill = FILL_FILL
    cdn = ws.cell(row=r2, column=4); cdn.number_format = 'dd mmm'; cdn.fill = FILL_FILL
    ws.cell(row=r2, column=5, value=n).font = F_NOTE
    for col in range(1, 6):
        ws.cell(row=r2, column=col).border = THIN_BOT
    r2 += 1

# ------------------------------------------------------- KPI Key Drivers (delivery-first model)
ws = wb.create_sheet('KPI Key Drivers')
set_widths(ws, {'A': 18, 'B': 52, 'C': 30, 'D': 30, 'E': 32})
title(ws, 'A1', 'KPI Key Drivers, per owner (delivery-first, per the Jul call with Ivan)')
note(ws, 'A2', 'Scored: delivery Speed and Quality against the seniority bar. Context: the line numbers every PM must know cold (management-owned, not punished on). Growth: initiative, assessed for development, not bonus.')
hdr_row(ws, 4, ['Owner', 'Context: lines and numbers to know cold', 'Scored: Speed', 'Scored: Quality', 'Growth: Initiative'])
SPEED_TXT = 'Agreed steps delivered on time or faster (auto-scored on the tracker)'
QUAL_TXT = 'Delivered at the level expected for the bar, thought through, minimal resyncs'
INIT_TXT = 'Connects the numbers to actions; brings prioritized initiatives (logged)'
drivers = [
    ('Anton Betia', 'Deals and Solar: monthly disbursement at positive NPM (staged ramp TBD), acceptance and conversion, reject-salvage plan. Think like a marketer: D2C, self-service, drive own inventory.'),
    ('Manzil Balani', 'Deals BD: managed accounts (ProTech, Maxicare, 3Cat), new merchants live and disbursing, ticket uplift (TBD). National coverage per partner, else regional deals.'),
    ('JV Manlangit', 'Motorcycles and SAP: Moto monthly disbursement (TBD, gated by Nhat), disbursement per active sales agent (TBD).'),
    ('Elias Marcella', 'Credit Line and Card: CL utilization and drawdown (set ~30 days post-launch), card unit economics (CAC target TBD). Senior bar: tasks arrive fully thought through.'),
    ('Jastin Lagumbay', 'QRPh and Merchant Onboarding: QRPh run rate (TBD, star line), onboarding cycle time, zero config errors.'),
    ('Nhat Phan', 'Unit economics and P&L: PBT-per-line model live monthly, pricing and repricing coverage, every launch gated pre-launch.'),
    ('Jenny Ho', 'Business Intelligence: metric-tree dashboard live and automated (Tier 1 to 3, one source), reporting turnaround.'),
    ('Billy Quiambao', 'IT Support and PVT: release quality gate (PVT pass rate, defects caught pre-prod), support TAT within SLA.'),
]
r = 5
for owner, ctx in drivers:
    ws.cell(row=r, column=1, value=owner).font = F_BODYB
    ws.cell(row=r, column=2, value=ctx).font = F_BODY
    ws.cell(row=r, column=3, value=SPEED_TXT).font = F_BODY
    ws.cell(row=r, column=4, value=QUAL_TXT).font = F_BODY
    ws.cell(row=r, column=5, value=INIT_TXT).font = F_BODY
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = THIN_BOT
        ws.cell(row=r, column=col).alignment = WRAP
    r += 1
r += 1
note(ws, f'A{r}', 'Targets and monthly actuals live on the KPI Targets H2 tab. Speed and Resyncs are captured per project on the tracker. Delivery scores and initiative log are kept by Kurt in the separate Delivery Scoreboard file.')
ws.freeze_panes = 'A5'

ORDER = ['README', 'Master Rollup', 'KPI Targets H2',
         'JV', 'Anton', 'Manzil', 'Elias', 'Jastin', 'KPI Key Drivers']
wb._sheets = [wb[n] for n in ORDER]
wb.save(OUT)
print('saved', OUT)
