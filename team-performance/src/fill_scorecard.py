#!/usr/bin/env python3
"""Fill the QuantumLight (Revolut) Product Owner scorecard as a worked example.

Illustrative Mid-level PM, no real person. Answers are chosen to demonstrate
the mechanics: sequential tier gating, the seniority talent bar, the
Deliverables = (Speed + Quality) x Complexity product, and how group grades
roll up to the overall grade.
"""
import openpyxl

SRC = 'scorecard.xlsx'
OUT = 'Scorecard_Filled_Example_Mid_PM.xlsx'

wb = openpyxl.load_workbook(SRC)
ws = wb['Assessment Scorecard']

ws['F4'] = 'Sample PM (illustrative, not a real person)'
ws['F5'] = 'Mid'

answers = {}
def setrange(rows, val):
    for r in rows:
        answers[r] = val

# Culture: Dream Team -> Intermediate (all POOR No, BASIC Yes, INTERMEDIATE Yes)
setrange(range(22, 26), 'No')      # POOR
setrange(range(26, 29), 'Yes')     # BASIC
setrange(range(29, 32), 'Yes')     # INTERMEDIATE
setrange(range(32, 39), 'No')      # ADVANCED + EXCEPTIONAL/EXPERT

# Culture: Push the Limits -> Basic (stops at BASIC; INTERMEDIATE No)
setrange(range(40, 45), 'No')      # POOR
setrange(range(45, 49), 'Yes')     # BASIC
setrange(range(49, 60), 'No')      # INTERMEDIATE and above

# Culture: Get It Done -> Intermediate
setrange(range(61, 64), 'No')      # POOR
setrange(range(64, 66), 'Yes')     # BASIC
setrange(range(66, 68), 'Yes')     # INTERMEDIATE
setrange(range(68, 71), 'No')      # ADVANCED + EXCEPTIONAL

# Deliverables: Complexity 4, Quality 4, Speed 4 -> 4 x (4+4) = 32 -> Advanced
setrange([73], 'No');  setrange([74, 75, 76], 'Yes'); setrange([77], 'No')   # Complexity
setrange([79], 'No');  setrange([80, 81, 82], 'Yes'); setrange([83], 'No')   # Quality
setrange([85], 'No');  setrange([86, 87, 88], 'Yes'); setrange([89], 'No')   # Speed

# Skills: Problem Solving -> Advanced
setrange([92], 'No'); setrange(range(93, 99), 'Yes'); setrange([99], 'No')
# Skills: Lifecycle Management -> Intermediate
setrange([101], 'No'); setrange(range(102, 108), 'Yes'); setrange(range(108, 113), 'No')
# Skills: Product Sense -> Intermediate
setrange(range(114, 117), 'No'); setrange(range(117, 127), 'Yes'); setrange(range(127, 135), 'No')

for r, v in answers.items():
    ws[f'M{r}'] = v

# KPI space (top right): the 3 KPIs from the tracker, placeholders until agreed
ws['M4'] = 'KPI - Q3 26'
ws['M5'] = 'D1 line outcome: monthly disbursement vs target (P__, TBD with Ivan)'
ws['Q5'] = 'TBD end Q3'
ws['R5'] = 'Tracker: KPI Targets H2'
ws['M6'] = 'D2 key driver: acceptance and conversion vs target (TBD)'
ws['Q6'] = 'TBD end Q3'
ws['R6'] = 'Tracker: KPI Targets H2'
ws['M7'] = 'D3 delivery: agreed roadmap steps shipped on pace (on-time rate)'
ws['Q7'] = 'On track'
ws['R7'] = 'Delivery Scoreboard'

# Feedback
ws['M12'] = 'Owns problems end to end, strong bias to action'
ws['M13'] = 'Data-driven prioritization, solid grasp of the metric tree'
ws['M14'] = 'Reliable delivery track record in the exec tracker'
ws['M16'] = 'Push the limits: propose new levers, not only execute the agreed ones'
ws['M17'] = 'Sharpen margin and risk diagnosis (Tier 2 questions)'
ws['M18'] = 'Raise first-pass quality to cut iterations to done'

# Evidence examples next to key statements (Examples column)
ws['Q30'] = 'e.g. led partner sync during launch week without being asked'
ws['Q31'] = 'e.g. flagged and closed the weakest vendor option'
ws['Q66'] = 'e.g. shipped the launch comms flow end to end'
ws['Q96'] = 'e.g. traced a conversion drop to one config error'

wb.save(OUT)
print('saved', OUT)
